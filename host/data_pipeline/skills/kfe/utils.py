"""
KFE (한국핵융합에너지연구원) 기관별 유틸리티
- 데이터 로드 (파일/URL/Excel)
- DataON 등록 폼 생성
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from shared.lib.dataon_reg import (
    DataON_연구데이터등록,
    기본정보, 인물, 인물구분,
    공개및라이선스설정, 파일데이터,
    국내외구분, 라이선스,
)
from shared.src.collectors.file_collector import FileCollector
from shared.src.collectors.api_collector import APICollector
from shared.src.transformers.cleaner import Cleaner
from shared.src.transformers.mapper import Mapper

# ──────────────────────────────────────────
# 기관 정보
# ──────────────────────────────────────────

INSTITUTION_INFO: dict[str, str] = {
    "기관명_한글": "한국핵융합에너지연구원",
    "기관명_영문": "Korea Institute of Fusion Energy",
    "약어": "KFE",
    "분야": "에너지·자원·원자력",
    "과학기술표준분류": "핵융합",
}

# ──────────────────────────────────────────
# 소스 필드 → DataON 필드 매핑
# 실제 KFE 데이터 포맷에 맞게 조정 필요
# ──────────────────────────────────────────

FIELD_MAP: dict[str, str] = {
    # 제목
    "title":              "제목_주언어",
    "title_ko":           "제목_주언어",
    "data_title":         "제목_주언어",
    "title_en":           "제목_부언어",
    # 설명
    "description":        "설명_주언어",
    "abstract":           "설명_주언어",
    "summary":            "설명_주언어",
    # 키워드
    "keyword":            "키워드_주언어",
    "keywords":           "키워드_주언어",
    "keyword_ko":         "키워드_주언어",
    "keyword_en":         "키워드_부언어",
    # 날짜
    "date":               "생성일자",
    "created_at":         "생성일자",
    "pub_date":           "생성일자",
    "year":               "생성일자",
    # 저자/기관
    "author":             "_저자",
    "creator":            "_저자",
    # 핵융합 특화
    "plasma_temperature": "_플라즈마온도",
    "magnetic_field":     "_자기장",
    "device":             "_장치명",
    "shot_number":        "_샷번호",
    "iter_task":          "_ITER과제",
    "discharge_type":     "_방전유형",
}


# ──────────────────────────────────────────
# 포맷 감지
# ──────────────────────────────────────────

def detect_format(source: str) -> str:
    s = source.strip()
    if s.startswith("http://") or s.startswith("https://"):
        return "url"
    return {
        ".csv":   "csv",
        ".json":  "json",
        ".jsonl": "jsonl",
        ".xlsx":  "excel",
        ".xls":   "excel",
    }.get(Path(s).suffix.lower(), "csv")


# ──────────────────────────────────────────
# 데이터 로드
# ──────────────────────────────────────────

def load_source(source: str) -> list[dict[str, Any]]:
    fmt = detect_format(source)
    if fmt == "url":
        collector = APICollector(config={"url": source})
        raw = collector.run()
        return raw if isinstance(raw, list) else [raw]
    if fmt == "excel":
        return _load_excel(source)
    collector = FileCollector(config={"path": source, "format": fmt})
    return collector.run()


def _load_excel(path: str) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("pip install openpyxl  # Excel 지원 필요")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [
        str(cell.value or f"col_{i}")
        for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
    ]
    return [
        {k: v for k, v in zip(headers, row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    ]


# ──────────────────────────────────────────
# DataON 폼 생성
# ──────────────────────────────────────────

def build_dataon_form(
    records: list[dict[str, Any]],
    source: str,
) -> DataON_연구데이터등록:
    cleaner = Cleaner(config={"strip_fields": ["title", "title_ko", "description", "abstract", "keyword", "author"]})
    mapper  = Mapper(config={"field_map": FIELD_MAP, "keep_extra": True})

    clean  = cleaner.run(records)
    mapped = mapper.run(clean)
    meta   = mapped[0] if mapped else {}

    return DataON_연구데이터등록(
        기본=기본정보(
            국내외=국내외구분.국내,
            제목_주언어=str(meta.get("제목_주언어") or "제목 미입력"),
            제목_부언어=meta.get("제목_부언어"),
            설명_주언어=str(meta.get("설명_주언어") or f"{INSTITUTION_INFO['기관명_한글']} 연구데이터"),
            키워드_주언어=_to_list(meta.get("키워드_주언어", [INSTITUTION_INFO["과학기술표준분류"]])),
            키워드_부언어=_to_list(meta.get("키워드_부언어", [])),
            과학기술표준분류=[INSTITUTION_INFO["과학기술표준분류"]],
            생성일자=str(meta["생성일자"]) if meta.get("생성일자") else None,
        ),
        인물정보=[
            인물(
                역할=인물구분.생성자,
                이름_주언어=str(meta["_저자"]) if meta.get("_저자") else None,
                기관_주언어=INSTITUTION_INFO["기관명_한글"],
                기관_부언어=INSTITUTION_INFO["기관명_영문"],
            )
        ],
        공개설정=공개및라이선스설정(
            라이선스종류=라이선스.저작자표시,
        ),
        파일=파일데이터(
            출처URL=[source] if detect_format(source) == "url" else [],
        ),
    )


def _to_list(val: Any) -> list[str]:
    if isinstance(val, list):
        return [str(v) for v in val if v is not None]
    if isinstance(val, str):
        return [v.strip() for v in val.replace(";", ",").split(",") if v.strip()]
    return [str(val)] if val is not None else []
