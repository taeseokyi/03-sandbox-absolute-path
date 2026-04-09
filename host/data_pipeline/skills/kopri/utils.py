"""
KOPRI (극지연구소) 기관별 유틸리티
- KPDC URL 스크래핑 또는 파일 로드 (scraper.py)
- KISTI LLM 번역·구조화 추출·저자 한글명 매핑 (translator.py)
- NTIS 과제정보 조회 및 연구분야 연동 (ntis_client.py)
- DataON 등록 폼 생성
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from data_pipeline.lib.dataon_reg import (
    DataON_연구데이터등록,
    기본정보, 인물, 인물구분,
    이메일, 이메일도메인,
    공개및라이선스설정, 파일데이터,
    추가정보, 수집기간, 수집지역, 수집지역유형,
    수집지역_Point, 수집지역_Box, 수집지역_Polygon, 좌표,
    국내외구분, 라이선스,
    연관정보, 과제정보, 관계유형_과제, 식별자유형_과제,
    과제책임자, 과제상세정보,
)
from .ntis_client import search_ntis_project, extract_project_fields
from data_pipeline.src.collectors.file_collector import FileCollector
from data_pipeline.src.collectors.api_collector import APICollector
from data_pipeline.src.transformers.cleaner import Cleaner
from data_pipeline.src.transformers.mapper import Mapper
from .scraper import scrape_kpdc_page
from .translator import (
    translate_en_to_ko,
    translate_keywords_en_to_ko,
    map_author_names,
    romanize_korean_name,
)

# ── 기관 정보 ─────────────────────────────
INSTITUTION_INFO: dict[str, str] = {
    "기관명_한글": "극지연구소",
    "기관명_영문": "Korea Polar Research Institute",
    "약어": "KOPRI",
    "분야": "기초·융합 과학",
    "과학기술표준분류": "해양학",
    "데이터포털": "KPDC (Korea Polar Data Center)",
}

_LICENSE_MAP: dict[str, 라이선스] = {
    "by":       라이선스.저작자표시,
    "by-nc":    라이선스.저작자표시_비영리,
    "by-nd":    라이선스.저작자표시_변경금지,
    "by-sa":    라이선스.저작자표시_동일조건변경허락,
    "by-nc-sa": 라이선스.저작자표시_비영리_동일조건,
    "by-nc-nd": 라이선스.저작자표시_비영리_변경금지,
}

FIELD_MAP: dict[str, str] = {
    "title":       "제목_부언어",
    "title_ko":    "제목_주언어",
    "title_en":    "제목_부언어",
    "description": "설명_부언어",
    "abstract":    "설명_부언어",
    "keyword":     "키워드_부언어",
    "keywords":    "키워드_부언어",
    "author":      "_저자",
    "date":        "생성일자",
    "latitude":    "_위도",
    "longitude":   "_경도",
    "lat":         "_위도",
    "lon":         "_경도",
    "start_date":  "_시작일",
    "end_date":    "_종료일",
    "doi":         "_doi",
    "project":     "_과제번호",
}


# ──────────────────────────────────────────
# 포맷 감지 / 데이터 로드
# ──────────────────────────────────────────

def _is_kpdc_url(source: str) -> bool:
    return "kpdc.kopri.re.kr" in source


def detect_format(source: str) -> str:
    s = source.strip()
    if s.startswith("http://") or s.startswith("https://"):
        return "url"
    return {
        ".csv":   "csv",
        ".json":  "json",
        ".jsonl": "jsonl",
        ".xlsx":  "excel",
        ".xls":   "excel",
    }.get(Path(s).suffix.lower(), "csv")


def load_source(source: str) -> list[dict[str, Any]]:
    """소스(KPDC URL / 파일)에서 메타데이터 로드"""
    if _is_kpdc_url(source):
        return [scrape_kpdc_page(source)]
    fmt = detect_format(source)
    if fmt == "excel":
        return _load_excel(source)
    if fmt == "url":
        collector = APICollector(config={"url": source})
        raw = collector.run()
        return raw if isinstance(raw, list) else [raw]
    collector = FileCollector(config={"path": source, "format": fmt})
    return collector.run()


def _load_excel(path: str) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("pip install openpyxl  # Excel 지원 필요")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [
        str(cell.value or f"col_{i}")
        for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
    ]
    return [
        {k: v for k, v in zip(headers, row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    ]


# ──────────────────────────────────────────
# DataON 폼 생성
# ──────────────────────────────────────────

def build_dataon_form(
    records: list[dict[str, Any]],
    source: str,
) -> DataON_연구데이터등록:
    """메타데이터 dict → DataON_연구데이터등록 폼

    처리 순서:
      1. 제목·설명·키워드·라이선스
      2. NTIS 조회 (인물 한글명 매핑을 위해 저자 구성 전에 선행)
      3. 저자 한글명 매핑 (NTIS Researchers 목록 활용)
      4. 인물목록 구성
      5. 과학기술표준분류 (NTIS researchArea 연동)
      6. 수집기간·수집지역·과제목록
    """
    meta = records[0] if records else {}

    # 파일 입력의 경우 Cleaner + Mapper 적용
    if not _is_kpdc_url(source):
        cleaner = Cleaner(config={"strip_fields": [
            "title", "title_ko", "description", "abstract", "keyword", "author"
        ]})
        mapper = Mapper(config={"field_map": FIELD_MAP, "keep_extra": True})
        meta = (mapper.run(cleaner.run(records)) or [{}])[0]

    # ── 제목 ───────────────────────────────
    title_en = str(meta.get("제목_부언어") or meta.get("제목_주언어") or "")
    title_ko = str(meta.get("제목_주언어") or "")
    if not title_ko or title_ko == title_en:
        title_ko = translate_en_to_ko(title_en, context="연구 데이터셋 제목") or title_en
    제목_주언어 = title_ko or "제목 미입력"
    제목_부언어 = title_en or None

    # ── 설명 ───────────────────────────────
    desc_en = str(meta.get("설명_부언어") or meta.get("설명_주언어") or "")
    desc_ko = str(meta.get("설명_주언어") or "")
    platform   = meta.get("_플랫폼", "")
    instrument = meta.get("_장비", "")
    if platform or instrument:
        parts = [f"Platform: {platform}" if platform else "",
                 f"Instrument: {instrument}" if instrument else ""]
        suffix = ". ".join(p for p in parts if p) + "."
        desc_en = f"{desc_en}\n{suffix}" if desc_en else suffix
    if not desc_ko or desc_ko == desc_en:
        desc_ko = translate_en_to_ko(desc_en, context="연구 데이터셋 설명") or desc_en
    설명_주언어 = desc_ko or f"{INSTITUTION_INFO['기관명_한글']} 연구데이터"
    설명_부언어 = desc_en or None

    # ── 키워드 ─────────────────────────────
    kw_en: list[str] = _to_list(meta.get("키워드_부언어", []))
    kw_ko: list[str] = _to_list(meta.get("키워드_주언어", []))
    if not kw_ko:
        kw_ko = translate_keywords_en_to_ko(kw_en) if kw_en else [INSTITUTION_INFO["과학기술표준분류"]]

    # ── 라이선스 ───────────────────────────
    cc_license = _LICENSE_MAP.get(str(meta.get("_license_key", "")).lower(), 라이선스.기타)

    # ── 저자 목록 ──────────────────────────
    authors: list[str] = meta.get("_저자목록", [])
    if not authors and meta.get("_저자"):
        authors = _to_list(meta["_저자"])
    emails_list: list[str] = meta.get("_이메일목록", [])

    # ── NTIS 조회 (인물목록 구성 전에 선행) ──
    ntis_data = None
    ntis_fields: dict[str, Any] = {}
    kpdc_과제번호 = meta.get("_과제번호")
    ntis_과제번호 = meta.get("_ntis_과제번호")
    과제명_영문   = meta.get("_과제명_영문")

    if ntis_과제번호:
        ntis_data = search_ntis_project(ntis_과제번호)
    if ntis_data:
        ntis_fields = extract_project_fields(ntis_data)

    # ── 저자 한글명 매핑 (NTIS Researchers 활용) ──
    name_map: dict[str, str] = {}
    if authors and ntis_data:
        ko_researchers = [
            r.strip()
            for r in ntis_data.get("Researchers", {}).get("Name", "").split(";")
            if r.strip()
        ]
        if ko_researchers:
            name_map = map_author_names(authors, ko_researchers)

    # ── 인물목록 ───────────────────────────
    인물목록: list[인물] = []
    for i, name in enumerate(authors):
        email_str = emails_list[i] if i < len(emails_list) else None
        인물목록.append(인물(
            역할=인물구분.생성자 if i == 0 else 인물구분.기여자,
            이름_주언어=name_map.get(name),
            이름_부언어=name,
            기관_주언어=INSTITUTION_INFO["기관명_한글"],
            기관_부언어=INSTITUTION_INFO["기관명_영문"],
            email=_parse_email(email_str) if email_str else None,
        ))
    if not 인물목록:
        인물목록 = [인물(
            역할=인물구분.생성자,
            기관_주언어=INSTITUTION_INFO["기관명_한글"],
            기관_부언어=INSTITUTION_INFO["기관명_영문"],
        )]

    # ── 과학기술표준분류 (NTIS researchArea 연동) ──
    ntis_area = ntis_fields.get("연구분야") or ""
    if ntis_area:
        base_area = re.sub(r"\(.*\)", "", ntis_area).strip()
        과학기술표준분류 = list(dict.fromkeys([INSTITUTION_INFO["과학기술표준분류"], base_area]))
    else:
        과학기술표준분류 = [INSTITUTION_INFO["과학기술표준분류"]]

    # ── 수집기간 ───────────────────────────
    기간목록 = []
    시작 = meta.get("_시작일")
    종료 = meta.get("_종료일")
    if 시작 or 종료:
        기간목록 = [수집기간(시작일자=시작, 종료일자=종료)]

    # ── 수집지역 (Point / Box / Polygon) ──────
    지역목록 = []
    lat = meta.get("_위도")
    lon = meta.get("_경도")
    coord_pairs: list[tuple] = meta.get("_좌표목록", [])
    geom_type = meta.get("_지역유형", "")

    if coord_pairs and geom_type == "Box":
        if len(coord_pairs) == 2:
            lats = sorted(c[0] for c in coord_pairs)
            lons = sorted(c[1] for c in coord_pairs)
            four = [
                좌표(위도=lats[0], 경도=lons[0]),
                좌표(위도=lats[0], 경도=lons[1]),
                좌표(위도=lats[1], 경도=lons[1]),
                좌표(위도=lats[1], 경도=lons[0]),
            ]
        else:
            four = [좌표(위도=c[0], 경도=c[1]) for c in coord_pairs[:4]]
        지역목록 = [수집지역(유형=수집지역유형.Box, 수집지역정보=수집지역_Box(좌표목록=four))]
    elif coord_pairs and geom_type == "Polygon":
        poly = [좌표(위도=c[0], 경도=c[1]) for c in coord_pairs]
        while len(poly) < 4:
            poly.append(poly[-1])
        지역목록 = [수집지역(유형=수집지역유형.Polygon, 수집지역정보=수집지역_Polygon(좌표목록=poly))]
    elif lat is not None and lon is not None:
        지역목록 = [수집지역(
            유형=수집지역유형.Point,
            수집지역정보=수집지역_Point(위도=float(lat), 경도=float(lon)),
        )]

    # ── 과제목록 ───────────────────────────
    과제목록 = []
    if ntis_data and ntis_fields:
        f = ntis_fields
        과제목록 = [과제정보(
            관계유형=관계유형_과제.유발과제,
            식별자유형=식별자유형_과제.NTIS,
            식별자=f["ntis_과제번호"],
            과제명_한글=f["과제명_한글"],
            과제명_영문=과제명_영문 or f["과제명_영문"] or None,
            부처명=f["부처명"],
            과제수행기관=f["과제수행기관"],
            과제책임자=과제책임자(
                책임자명_한글=f["책임자명_한글"],
                책임자명_영문=romanize_korean_name(f["책임자명_한글"]),
                기관명_한글=f["과제수행기관"],
            ) if f["책임자명_한글"] else None,
            상세입력=과제상세정보(
                과제관리기관=f["과제관리기관"],
                기준년도=f["기준년도"],
                연구기간_시작일=f["연구기간_시작"],
                연구기간_종료일=f["연구기간_종료"],
                총연구비=f["총연구비"],
                키워드_한글=f["키워드_한글"],
                키워드_영문=f["키워드_영문"],
            ) if any([f["과제관리기관"], f["키워드_한글"]]) else None,
        )]
    elif kpdc_과제번호 or ntis_과제번호:
        # NTIS 조회 실패 시 fallback
        과제명_한글 = None
        if 과제명_영문:
            과제명_한글 = translate_en_to_ko(과제명_영문, context="연구 과제명") or None
        식별자 = ntis_과제번호 or kpdc_과제번호
        유형 = 식별자유형_과제.NTIS if ntis_과제번호 else 식별자유형_과제.ETC
        과제목록 = [과제정보(
            관계유형=관계유형_과제.유발과제,
            식별자유형=유형,
            식별자=식별자,
            과제명_한글=과제명_한글,
            과제명_영문=과제명_영문,
        )]

    # ── 출처URL ────────────────────────────
    출처URL = [source] if source.startswith("http") else []
    doi = meta.get("_doi")
    if doi and doi not in 출처URL:
        출처URL.append(doi)

    return DataON_연구데이터등록(
        연관=연관정보(과제목록=과제목록),
        기본=기본정보(
            국내외=국내외구분.국내,
            제목_주언어=제목_주언어,
            제목_부언어=제목_부언어,
            설명_주언어=설명_주언어,
            설명_부언어=설명_부언어,
            키워드_주언어=kw_ko,
            키워드_부언어=kw_en,
            과학기술표준분류=과학기술표준분류,
            생성일자=meta.get("생성일자"),
        ),
        인물정보=인물목록,
        추가=추가정보(
            데이터수집기간=기간목록,
            데이터수집지역=지역목록,
        ),
        공개설정=공개및라이선스설정(라이선스종류=cc_license),
        파일=파일데이터(출처URL=출처URL),
    )


# ──────────────────────────────────────────
# 내부 헬퍼
# ──────────────────────────────────────────

def _parse_email(email_str: str) -> 이메일:
    if "@" in email_str:
        local, domain_str = email_str.split("@", 1)
        try:
            return 이메일(id=local, domain=이메일도메인(domain_str))
        except ValueError:
            return 이메일(id=email_str, domain=이메일도메인.직접입력)
    return 이메일(id=email_str, domain=이메일도메인.직접입력)


def _to_list(val: Any) -> list[str]:
    if isinstance(val, list):
        return [str(v) for v in val if v is not None]
    if isinstance(val, str):
        return [v.strip() for v in val.replace(";", ",").split(",") if v.strip()]
    return [str(val)] if val is not None else []
